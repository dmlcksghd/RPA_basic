from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])   # A, B, C
for i in range(1, 11):  # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 컬럼만 가져오기
print(col_B)    # (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>, <Cell 'Sheet'.B3>, <Cell 'Sheet'.B4>, <Cell 'Sheet'.B5>, <Cell 'Sheet'.B6>, <Cell 'Sheet'.B7>, <Cell 'Sheet'.B8>, <Cell 'Sheet'.B9>, <Cell 'Sheet'.B10>, <Cell 'Sheet'.B11>)
for cell in col_B:
    print(cell.value)

# 컬럼 여러개 가져오기
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]   # 1번째 row만 가져오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range2 = ws[2:ws.max_row]
for rows in row_range2:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # 셀의 좌표 정보
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end=" ")   # A
        print(xy[1], end=" ")   # 1
    print()

# 전체 rows
print(tuple(ws.rows))


wb.save("sample.xlsx")