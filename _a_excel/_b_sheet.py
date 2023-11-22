from openpyxl import Workbook
wb = Workbook()
# wb.avtive # 활성화 된 Sheet 가져오기
ws = wb.create_sheet()  # 새로운 Sheet 생성
ws.title = "MySheet"    # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"   # RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)    # 2번째 index 에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근
# new_ws.

print(wb.sheetnames)    # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "copy_A1_new"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")